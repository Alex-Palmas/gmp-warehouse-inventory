#!/usr/bin/env python3
"""Generate companion Excel reporting template. Not a Part 11 system of record."""
from openpyxl import Workbook
from openpyxl.styles import Font, Fill, PatternFill, Alignment, Border, Side, Protection
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import FormulaRule
from openpyxl.chart import BarChart, Reference
from datetime import datetime, timezone
from pathlib import Path

OUT = Path("/workspace/gmp-warehouse-inventory/exports/GMP_Warehouse_Inventory_Register_v1.1.xlsx")
DOC = "DOC-WH-INV-001"
VER = "1.1"
BANNER = "Not validated — do not use for GMP decisions until IQ/OQ/PQ approved."
NAVY = "1A365D"
NAVY_FILL = PatternFill("solid", fgColor=NAVY)
AMBER = PatternFill("solid", fgColor="FEFCBF")
WHITE = Font(color="FFFFFF", bold=True, name="Calibri", size=10)
HEADER_FONT = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
TITLE = Font(name="Calibri", size=16, bold=True, color=NAVY)
WARN = Font(name="Calibri", size=12, bold=True, color="C05621")
THIN = Border(
    left=Side(style="thin", color="CBD5E0"),
    right=Side(style="thin", color="CBD5E0"),
    top=Side(style="thin", color="CBD5E0"),
    bottom=Side(style="thin", color="CBD5E0"),
)
WRAP = Alignment(wrap_text=True, vertical="center")

ITEM_TYPES = [
    "Raw Material", "Excipient", "API", "Intermediate", "Packaging Component",
    "Finished Product", "Sample", "Retain Sample", "Reference Standard", "Consumable",
]
STATUSES = ["Quarantine", "Released", "Rejected", "Restricted", "Hold", "Issued", "Consumed", "Destroyed"]
UOMS = ["kg", "g", "mg", "L", "mL", "each", "bottle", "drum", "bag", "vial", "pack"]
STORAGE = [
    "CRT 15–25 °C", "2–8 °C", "−20 °C", "−80 °C",
    "Controlled humidity", "Light-sensitive", "Flammable",
]
PHARMA = ["USP", "EP", "JP", "In-house", "NF", "BP"]
CONTAINERS = ["Drum", "Bag", "Bottle", "Vial", "Carton", "Pallet", "IBC", "Ampoule", "Blister", "Other"]
ROLES = ["System Administrator", "Warehouse Supervisor", "Warehouse Operator", "QA", "QC", "Read-Only"]
YN = ["Y", "N"]
QA_DISP = ["Release", "Reject", "Restricted"]


def style_header(ws, ncols, freeze=True):
    for c in range(1, ncols + 1):
        cell = ws.cell(1, c)
        cell.fill = NAVY_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(wrap_text=True, vertical="center")
        cell.border = THIN
    ws.row_dimensions[1].height = 28
    if freeze:
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{get_column_letter(ncols)}1"
    ws.auto_filter.ref = f"A1:{get_column_letter(ncols)}{max(ws.max_row, 2)}"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_setup.paperSize = ws.PAPERSIZE_TABLOID
    ws.oddFooter.left.text = f"{DOC} v{VER} | {BANNER}"
    ws.oddFooter.right.text = "REPORTING TEMPLATE — not the system of record"


def footer_rows(ws, start_row, ncols):
    ws.cell(start_row, 1, f"{DOC} v{VER} | {BANNER}")
    ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=min(ncols, 8))
    ws.cell(start_row, 1).font = Font(italic=True, size=9, color="744210")
    r2 = start_row + 1
    ts = datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")
    ws.cell(r2, 1, f"Generated {ts} | App 1.1.0 | REPORTING TEMPLATE / offline backup format — not a 21 CFR Part 11 system")
    ws.merge_cells(start_row=r2, start_column=1, end_row=r2, end_column=min(ncols, 8))
    ws.cell(r2, 1).font = Font(italic=True, size=9, color="4A5568")


def autosize(ws, widths=None):
    if widths:
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
        return
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        maxlen = 10
        for cell in col[:30]:
            if cell.value:
                maxlen = min(max(maxlen, len(str(cell.value)) + 2), 42)
        ws.column_dimensions[letter].width = maxlen


def dv(ws, formula, cells):
    v = DataValidation(type="list", formula1=formula, allow_blank=True)
    v.error = "Select from controlled list"
    v.errorTitle = "Controlled vocabulary"
    v.showErrorMessage = True
    ws.add_data_validation(v)
    v.add(cells)


wb = Workbook()

# ----- Cover -----
cv = wb.active
cv.title = "Cover"
cv.sheet_properties.tabColor = "C05621"
cv["A1"] = "GMP WAREHOUSE INVENTORY REGISTER"
cv["A1"].font = TITLE
cv.merge_cells("A1:F1")
cv["A3"] = "CONTROLLED DOCUMENT — REPORTING TEMPLATE / OFFLINE BACKUP FORMAT"
cv["A3"].font = Font(bold=True, size=12, color="C53030")
cv.merge_cells("A3:F3")
cv["A5"] = BANNER
cv["A5"].font = WARN
cv.merge_cells("A5:F5")
cv["A5"].fill = AMBER

cover_rows = [
    (7, "Document number", DOC),
    (8, "Version", VER),
    (9, "Application", "DOC-WH-INV-001 v1.1 (IndexedDB system of record lives in the web app, not this file)"),
    (10, "Site", "[SITE]"),
    (11, "Owner", "[OWNER]"),
    (12, "Date", "[DATE]"),
    (13, "Classification", "GxP-relevant template — not approved, not validated"),
]
for r, k, v in cover_rows:
    cv.cell(r, 1, k).font = Font(bold=True)
    cv.cell(r, 2, v)
    cv.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)

cv["A15"] = "INTENDED USE"
cv["A15"].font = Font(bold=True, color=NAVY, size=12)
cv["A16"] = (
    "This workbook mirrors the register structure of the warehouse inventory application. "
    "It may be used as a print/archive layout, data-migration mapping aid, or offline recording sheet "
    "during dual-running. It is NOT the system of record. It is NOT a 21 CFR Part 11 electronic record system. "
    "Dropdowns are uncontrolled once the file is copied. Anyone with the file can edit cells. "
    "Do not use this file for GMP release, issue, or destruction decisions."
)
cv["A16"].alignment = Alignment(wrap_text=True, vertical="top")
cv.merge_cells("A16:F20")
cv.row_dimensions[16].height = 48

cv["A22"] = "SHEETS"
cv["A22"].font = Font(bold=True, color=NAVY, size=12)
sheets_info = [
    "Cover — this page",
    "Instructions — how to fill if used offline",
    "Lookups — controlled lists (source of data-validation lists)",
    "Material Master — seed materials",
    "Inventory Register — seed containers (14 rows) with dropdowns",
    "Goods Receipt log — blank log",
    "Movements — blank + example",
    "QA Dispositions — blank + example",
    "Audit Trail — headers + example rows (append-only conceptually; Excel cannot enforce)",
    "User Access Log — headers + example",
]
for i, s in enumerate(sheets_info):
    cv.cell(23 + i, 1, f"{i+1}. {s}")

cv["A35"] = "VALIDATION STATUS"
cv["A35"].fill = AMBER
cv["A35"].font = Font(bold=True)
cv.merge_cells("A35:F35")
cv["A36"] = BANNER
cv.merge_cells("A36:F36")
cv["A38"] = f"{DOC} v{VER} | Generated as a companion to the SPA | Footer required on all prints"
cv["A38"].font = Font(italic=True, size=9)
autosize(cv, [28, 28, 22, 22, 22, 22])
cv.oddFooter.left.text = f"{DOC} v{VER}"
cv.oddFooter.center.text = BANNER

# ----- Instructions -----
ins = wb.create_sheet("Instructions")
ins["A1"] = "Instructions (offline use only)"
ins["A1"].font = TITLE
ins.merge_cells("A1:B1")
steps = [
    "1. The web application is the system of record. Prefer it.",
    "2. If this file is used during dual-running, print Cover + Register; do not rely on file sharing as an audit trail.",
    "3. Material codes MUST exist on Material Master before a receipt line is added.",
    "4. Inventory serial format is WH-YYYY-NNNNNN. Never reuse a serial, including cancelled drafts in the app (app allocates only on submit). If recording on paper/excel, the site serial log is the allocator — do not invent numbers.",
    "5. Receipt status is always Quarantine. QA completes QA Dispositions sheet (or the app form) before Release.",
    "6. Do not issue unless Status=Released and Expiry >= today. FEFO: issue earliest expiry released lot of that material first.",
    "7. No hard-delete. Mark Issued / Consumed / Destroyed.",
    "8. Corrections: strike-through on paper; in app use forms with reason for change. Do not overwrite Excel history if this file is archived — add a new row and note the correction.",
    "9. Lookups sheet drives dropdowns. Do not rename those columns.",
    "10. Demo seed data is fictional. Replace with [SITE] data only under change control after PQ.",
    "11. Document control: every print must show DOC-WH-INV-001 v1.1 and the not-validated banner until QA removes it via change control.",
]
for i, s in enumerate(steps, start=3):
    ins.cell(i, 1, s)
    ins.merge_cells(start_row=i, start_column=1, end_row=i, end_column=2)
    ins.cell(i, 1).alignment = WRAP
    ins.row_dimensions[i].height = 32
ins.column_dimensions["A"].width = 120
footer_rows(ins, 16, 2)

# ----- Lookups -----
lk = wb.create_sheet("Lookups")
lk_headers = ["ItemType", "Status", "UOM", "StorageCondition", "Pharmacopeia", "ContainerType", "Role", "YN", "QADisposition"]
lk.append(lk_headers)
style_header(lk, len(lk_headers))
lists = [ITEM_TYPES, STATUSES, UOMS, STORAGE, PHARMA, CONTAINERS, ROLES, YN, QA_DISP]
max_len = max(len(x) for x in lists)
for i in range(max_len):
    row = []
    for lst in lists:
        row.append(lst[i] if i < len(lst) else None)
    lk.append(row)
# named ranges via tables for DV formulas
# We'll use sheet references Lookups!$A$2:$A$11 etc.
autosize(lk)
footer_rows(lk, max_len + 4, len(lk_headers))
# keep a marker so DV ranges stay stable — pad
lk["A30"] = "Controlled lists — do not insert rows above row 2."
lk["A30"].font = Font(italic=True, size=9)

# ----- Material Master -----
mm = wb.create_sheet("Material Master")
mm_h = ["MaterialCode", "MaterialName", "ItemType", "GradeSpec", "Pharmacopeia", "DefaultUOM", "DefaultStorage", "SamplingRequired", "Active", "CreatedBy", "CreatedOnUTC"]
mm.append(mm_h)
materials = [
    ("RM-001", "Lactose Monohydrate", "Excipient", "NF / Ph. Eur. 200 mesh", "USP", "kg", "CRT 15–25 °C", "Y", "Y"),
    ("RM-002", "Microcrystalline Cellulose", "Excipient", "PH-102", "USP", "kg", "CRT 15–25 °C", "Y", "Y"),
    ("RM-003", "Magnesium Stearate", "Excipient", "Vegetable grade", "NF", "kg", "CRT 15–25 °C", "Y", "Y"),
    ("API-001", "Ibuprofen", "API", "USP micronized", "USP", "kg", "CRT 15–25 °C", "Y", "Y"),
    ("API-002", "Acetaminophen", "API", "USP compactable", "USP", "kg", "CRT 15–25 °C", "Y", "Y"),
    ("PKG-001", "HDPE Bottle 100 mL", "Packaging Component", "Pharma grade, 38-400 neck", "In-house", "each", "CRT 15–25 °C", "N", "Y"),
    ("PKG-002", "Child-Resistant Cap 38 mm", "Packaging Component", "CRC, induction liner", "In-house", "each", "CRT 15–25 °C", "N", "Y"),
    ("INT-001", "Granulation Blend A", "Intermediate", "In-process spec IPS-A-01", "In-house", "kg", "Controlled humidity", "Y", "Y"),
    ("FP-001", "Ibuprofen Tablet 200 mg", "Finished Product", "NDA-spec, 200 mg", "USP", "bottle", "CRT 15–25 °C", "Y", "Y"),
    ("CON-001", "Nitrile Gloves L", "Consumable", "Powder-free exam", "In-house", "pack", "CRT 15–25 °C", "N", "Y"),
    ("RS-001", "Ibuprofen Reference Standard", "Reference Standard", "USP RS", "USP", "mg", "2–8 °C", "N", "Y"),
    ("SAM-001", "Retain Sample Vial 20 mL", "Retain Sample", "Amber glass", "In-house", "vial", "CRT 15–25 °C", "N", "Y"),
]
utc = "2026-01-15T16:00:00.000Z"
for m in materials:
    mm.append(list(m) + ["admin", utc])
style_header(mm, len(mm_h))
dv(mm, "Lookups!$A$2:$A$11", "C2:C200")
dv(mm, "Lookups!$E$2:$E$7", "E2:E200")
dv(mm, "Lookups!$C$2:$C$12", "F2:F200")
dv(mm, "Lookups!$D$2:$D$8", "G2:G200")
dv(mm, "Lookups!$H$2:$H$3", "H2:I200")
autosize(mm)
footer_rows(mm, len(materials) + 4, len(mm_h))

# ----- Inventory Register -----
inv = wb.create_sheet("Inventory Register")
inv_h = [
    "Serial", "Barcode", "MaterialCode", "MaterialName", "ItemType", "GradeSpec", "Pharmacopeia",
    "Manufacturer", "ManufacturerLot", "Supplier", "SupplierLot", "PODeliveryNote", "CoANumber",
    "InternalLot", "QtyReceived", "CurrentQty", "UOM", "NumberOfContainers", "ContainerType",
    "DateOfManufacture", "ReceiptDate", "ExpiryDate", "RetestDate",
    "Site", "Building", "Room", "Rack", "Shelf", "Bin",
    "StorageCondition", "Status", "SamplingRequired", "LinkedSampleIDs", "Comments",
    "CreatedBy", "CreatedOnUTC", "ModifiedBy", "ModifiedOnUTC",
    "QADisposition", "QASignedBy", "QAUserID", "QASignedAtUTC", "MeaningOfSignature",
]
inv.append(inv_h)

def loc(bin_):
    return ("MAIN", "WH-1", "RM-W01", "R" + bin_[0], "S1", bin_)

# 14 seed rows matching app
seed_inv = [
    # serial, code, name, type, status, qty, uom, expiry, mfrlot, bin, extra
    ("WH-2026-000001", "API-001", "Ibuprofen", "API", "Released", 25, 25, "kg", "2028-06-30", "MFR-IBU-4412", "Q1", "Release"),
    ("WH-2026-000002", "API-001", "Ibuprofen", "API", "Released", 25, 25, "kg", "2027-03-15", "MFR-IBU-3988", "Q1", "Release"),
    ("WH-2026-000003", "API-001", "Ibuprofen", "API", "Quarantine", 25, 25, "kg", "2029-01-10", "MFR-IBU-5100", "Q2", ""),
    ("WH-2026-000004", "RM-001", "Lactose Monohydrate", "Excipient", "Quarantine", 50, 50, "kg", "2028-11-01", "LAC-9921", "Q3", ""),
    ("WH-2026-000005", "RM-002", "Microcrystalline Cellulose", "Excipient", "Released", 40, 40, "kg", "2027-08-20", "MCC-2201", "Q4", "Release"),
    ("WH-2026-000006", "RM-003", "Magnesium Stearate", "Excipient", "Quarantine", 10, 10, "kg", "2027-12-01", "MGS-1104", "Q5", ""),
    ("WH-2026-000007", "API-002", "Acetaminophen", "API", "Released", 30, 30, "kg", "2025-12-31", "APAP-771", "Q6", "Release"),  # expired
    ("WH-2026-000008", "PKG-001", "HDPE Bottle 100 mL", "Packaging Component", "Released", 5000, 5000, "each", "2029-05-01", "BTL-330", "Q7", "Release"),
    ("WH-2026-000009", "PKG-002", "Child-Resistant Cap 38 mm", "Packaging Component", "Quarantine", 5000, 5000, "each", "2029-05-01", "CAP-330", "Q8", ""),
    ("WH-2026-000010", "INT-001", "Granulation Blend A", "Intermediate", "Restricted", 80, 80, "kg", "2026-12-15", "INTA-09", "Q9", "Restricted"),
    ("WH-2026-000011", "FP-001", "Ibuprofen Tablet 200 mg", "Finished Product", "Quarantine", 200, 200, "bottle", "2028-02-28", "LOT-FP-1001", "QA", ""),
    ("WH-2026-000012", "RS-001", "Ibuprofen Reference Standard", "Reference Standard", "Released", 100, 100, "mg", "2026-10-01", "USP-RS-IBU", "QB", "Release"),
    ("WH-2026-000013", "CON-001", "Nitrile Gloves L", "Consumable", "Released", 40, 40, "pack", "2028-01-01", "GLV-55", "QC", "Release"),
    ("WH-2026-000014", "API-001", "Ibuprofen", "API", "Issued", 25, 0, "kg", "2027-01-20", "MFR-IBU-3001", "QD", "Release"),
]
iutc = "2026-02-01T17:00:00.000Z"
for s in seed_inv:
    serial, code, name, typ, status, qrec, qcur, uom, exp, mfr, bin_, qad = s
    site, bld, room, rack, shelf, bn = loc(bin_)
    comments = "SEED: expired lot for OQ expiry-block test" if exp == "2025-12-31" else "Seed data"
    sampling = "Y" if typ in ("API", "Excipient", "Raw Material") else "N"
    qa_name = "Jordan QA" if qad else ""
    qa_id = "qa" if qad else ""
    qa_at = iutc if qad else ""
    meaning = ""
    if qad == "Release":
        meaning = "I attest this lot is released for GMP use per CoA and specification."
    elif qad == "Restricted":
        meaning = "Restricted to engineering use only pending investigation."
    ctype = "Carton" if uom in ("each", "bottle") else "Drum"
    storage = "2–8 °C" if code == "RS-001" else "CRT 15–25 °C"
    row = [
        serial, serial, code, name, typ, "See material master", "USP",
        "Demo Manufacturer Inc.", mfr, "Demo Supplier LLC", f"SUP-{mfr}", f"PO-2026-{serial[-4:]}", f"COA-{mfr}",
        f"IL-{serial[-6:]}", qrec, qcur, uom, 1, ctype,
        "2025-01-15", "2026-02-01", exp, "",
        site, bld, room, rack, shelf, bn,
        storage, status, sampling, "", comments,
        "admin", iutc, "qa" if qad else "admin", iutc,
        qad, qa_name, qa_id, qa_at, meaning,
    ]
    inv.append(row)

style_header(inv, len(inv_h))
# data validations on seed+blank rows
dv(inv, "Lookups!$A$2:$A$11", "E2:E200")
dv(inv, "Lookups!$E$2:$E$7", "G2:G200")
dv(inv, "Lookups!$C$2:$C$12", "Q2:Q200")
dv(inv, "Lookups!$F$2:$F$11", "S2:S200")
dv(inv, "Lookups!$D$2:$D$8", "AD2:AD200")
dv(inv, "Lookups!$B$2:$B$9", "AE2:AE200")
dv(inv, "Lookups!$H$2:$H$3", "AF2:AF200")
dv(inv, "Lookups!$I$2:$I$4", "AM2:AM200")

# highlight expired
from openpyxl.formatting.rule import CellIsRule
inv.conditional_formatting.add(
    "V2:V200",
    CellIsRule(operator="lessThan", formula=["TODAY()"], fill=PatternFill("solid", fgColor="FED7D7")),
)
# highlight quarantine
from openpyxl.formatting.rule import FormulaRule as FR
inv.conditional_formatting.add(
    "AE2:AE200",
    FR(formula=['$AE2="Quarantine"'], fill=PatternFill("solid", fgColor="FEEBC8")),
)

autosize(inv)
inv.column_dimensions["A"].width = 18
footer_rows(inv, len(seed_inv) + 4, len(inv_h))

# ----- Goods Receipt log -----
gr = wb.create_sheet("Goods Receipt log")
gr_h = ["ReceiptID", "Date", "SerialAllocated", "MaterialCode", "Qty", "UOM", "MfrLot", "InternalLot", "CoA", "PO", "LocationBin", "ReceivedBy", "StatusDefault", "Comments"]
gr.append(gr_h)
gr.append(["(example)", "2026-02-01", "WH-2026-000004", "RM-001", 50, "kg", "LAC-9921", "IL-000004", "COA-LAC-9921", "PO-2026-0004", "Q3", "admin", "Quarantine", "Seed example — do not treat as live"])
style_header(gr, len(gr_h))
dv(gr, "Lookups!$C$2:$C$12", "F2:F200")
autosize(gr)
footer_rows(gr, 6, len(gr_h))
gr["A8"] = "Serial is allocated only on successful submit in the application. Do not pre-assign from this sheet."

# ----- Movements -----
mv = wb.create_sheet("Movements")
mv_h = ["MovementID", "Serial", "Action", "Qty", "FromLocation", "ToLocation", "PerformedBy", "PerformedOnUTC", "Reason", "Comments"]
mv.append(mv_h)
mv.append(["MOV-example-1", "WH-2026-000014", "ISSUE", 25, "MAIN / WH-1 / RM-W01 / RQ / S1 / QD", "Batch BATCH-DEMO-1", "wh", iutc, "Dispensed to granulation", "Seed example"])
mv.append(["MOV-example-2", "WH-2026-000001", "RECEIVE", 25, "", "MAIN / WH-1 / RM-W01 / RQ / S1 / Q1", "admin", iutc, "Goods receipt", "Seed example"])
style_header(mv, len(mv_h))
autosize(mv)
footer_rows(mv, 8, len(mv_h))

# ----- QA Dispositions -----
qa = wb.create_sheet("QA Dispositions")
qa_h = ["DispositionID", "Serial", "Disposition", "ResultingStatus", "Reason", "SignedByPrintedName", "UserID", "SignedAtUTC", "MeaningOfSignature", "PasswordReentered"]
qa.append(qa_h)
qa.append([
    "QD-example-1", "WH-2026-000001", "Release", "Released",
    "CoA COA-MFR-IBU-4412 conforms; LIMS sample SAM-4412 pass",
    "Jordan QA", "qa", iutc,
    "I attest this container meets specification and is Released for GMP use.",
    "YES (in app only — never store the password in Excel)",
])
style_header(qa, len(qa_h))
dv(qa, "Lookups!$I$2:$I$4", "C2:C200")
autosize(qa)
footer_rows(qa, 6, len(qa_h))
qa["A8"] = "Excel cannot capture a compliant e-sign. Use the application e-sign modal. This sheet is a print image only."

# ----- Audit Trail -----
au = wb.create_sheet("Audit Trail")
au_h = ["AuditID", "TimestampUTC", "TimestampLocal", "UserID", "UserName", "Action", "RecordID", "Field", "OldValue", "NewValue", "ReasonForChange", "MeaningOfSignature"]
au.append(au_h)
au.append(["AUD-example-1", iutc, "02/01/2026, 09:00:00 PST", "admin", "Alex Supervisor", "RECEIVE", "WH-2026-000001", "status", "", "Quarantine", "Goods receipt created in Quarantine", ""])
au.append(["AUD-example-2", iutc, "02/01/2026, 09:10:00 PST", "qa", "Jordan QA", "QA_DISPOSITION", "WH-2026-000001", "status", "Quarantine", "Released", "CoA review complete", "I attest this container meets specification and is Released for GMP use."])
au.append(["AUD-example-3", iutc, "02/01/2026, 09:20:00 PST", "wh", "Sam Operator", "ISSUE", "WH-2026-000014", "currentQty", "25", "0", "Dispensed to granulation", ""])
style_header(au, len(au_h))
autosize(au)
footer_rows(au, 8, len(au_h))
au["A10"] = "APPEND-ONLY conceptually. Excel provides no immutability. The application audit store has no update/delete API. Do not edit example rows in an archived copy; add files instead."

# ----- User Access Log -----
ac = wb.create_sheet("User Access Log")
ac_h = ["AccessID", "TimestampUTC", "UserID", "UserName", "Event", "Detail"]
ac.append(ac_h)
ac.append(["ACC-example-1", iutc, "admin", "Alex Supervisor", "LOGIN", "role=Warehouse Supervisor"])
ac.append(["ACC-example-2", iutc, "qa", "Jordan QA", "LOGIN", "role=QA"])
ac.append(["ACC-example-3", iutc, "unknown", "", "LOGIN_FAIL", "Unknown or inactive user"])
style_header(ac, len(ac_h))
autosize(ac)
footer_rows(ac, 8, len(ac_h))


# ----- Roles -----
rl = wb.create_sheet("Roles")
rl_h = ["RoleID", "Name", "Description", "System", "Active"]
rl.append(rl_h)
roles_seed = [
    ("sysadmin", "System Administrator", "Access control; not a warehouse/QA actor (SoD).", "Y", "Y"),
    ("supervisor", "Warehouse Supervisor", "Warehouse ops, hold, user admin (cannot edit matrix).", "Y", "Y"),
    ("operator", "Warehouse Operator", "Receive, transfer, issue, return, cycle count, labels.", "Y", "Y"),
    ("qa", "QA", "Disposition, destroy, hold. Does not receive/issue/transfer (SoD).", "Y", "Y"),
    ("qc", "QC", "View, scan, cycle count, reprint.", "Y", "Y"),
    ("readonly", "Read-Only", "View dashboard, register, scan, audit.", "Y", "Y"),
]
for r in roles_seed:
    rl.append(list(r))
style_header(rl, len(rl_h))
autosize(rl)
footer_rows(rl, len(roles_seed) + 4, len(rl_h))

# ----- User Access Matrix -----
mx = wb.create_sheet("User Access Matrix")
caps = [
    "viewDashboard", "viewRegister", "viewAudit", "viewAccessLog", "scanLookup",
    "receive", "transfer", "issue", "returnToStock", "cycleCount", "reprintLabel",
    "hold", "qaDisposition", "destroy", "eSign",
    "adminMaterials", "adminUsers", "editPermissionMatrix", "backupRestore", "exportReports",
]
role_ids = ["sysadmin", "supervisor", "operator", "qa", "qc", "readonly"]
defaults = {
    "sysadmin": {"viewDashboard","viewRegister","viewAudit","viewAccessLog","scanLookup","adminUsers","editPermissionMatrix","exportReports","backupRestore"},
    "supervisor": {"viewDashboard","viewRegister","viewAudit","viewAccessLog","scanLookup","receive","transfer","issue","returnToStock","cycleCount","reprintLabel","hold","adminMaterials","adminUsers","exportReports","backupRestore"},
    "operator": {"viewDashboard","viewRegister","scanLookup","receive","transfer","issue","returnToStock","cycleCount","reprintLabel"},
    "qa": {"viewDashboard","viewRegister","viewAudit","viewAccessLog","scanLookup","qaDisposition","destroy","hold","adminMaterials","exportReports","backupRestore","eSign","reprintLabel"},
    "qc": {"viewDashboard","viewRegister","scanLookup","cycleCount","reprintLabel"},
    "readonly": {"viewDashboard","viewRegister","scanLookup","viewAudit"},
}
mx.append(["Capability"] + [r[1] for r in roles_seed])
for cap in caps:
    mx.append([cap] + ["Y" if cap in defaults[rid] else "N" for rid in role_ids])
style_header(mx, 7)
autosize(mx)
footer_rows(mx, len(caps) + 4, 7)
mx["A28"] = "SoD ON: qaDisposition XOR receive; destroy requires eSign; editPermissionMatrix XOR qaDisposition. Matrix v1 seeded."

# ----- User Access List -----
ual = wb.create_sheet("User Access List")
ual_h = ["UserID", "FullName", "RoleID", "Active", "Locked", "MustChangePassword", "Algorithm"]
ual.append(ual_h)
for uid, name, role in [
    ("sysadmin", "Casey SysAdmin", "sysadmin"),
    ("admin", "Alex Supervisor", "supervisor"),
    ("qa", "Jordan QA", "qa"),
    ("qc", "Morgan QC", "qc"),
    ("wh", "Sam Operator", "operator"),
    ("ro", "Riley ReadOnly", "readonly"),
]:
    ual.append([uid, name, role, "Y", "N", "Y", "sha256-salt (upgrades to pbkdf2-sha256 on login)"])
style_header(ual, len(ual_h))
autosize(ual)
footer_rows(ual, 10, len(ual_h))

for ws in (mm, inv, gr, mv, qa, au, ac, rl, mx, ual):
    if ws.max_row >= 1 and ws.max_column >= 1:
        ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{max(ws.max_row, 2)}"

OUT.parent.mkdir(parents=True, exist_ok=True)
wb.save(OUT)
print("wrote", OUT, "sheets", wb.sheetnames)
